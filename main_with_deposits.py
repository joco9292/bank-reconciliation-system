import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), 'processors'))
sys.path.append(os.path.join(os.path.dirname(__file__), 'matchers'))
import argparse
import pandas as pd
from datetime import datetime

def run_card_matching(card_summary_path: str, bank_statement_path: str, 
                      output_dir: str = '.', verbose: bool = False):
    """
    Run the credit card matching process.
    """
    from preprocess_bank_statement import preprocess_bank_statement
    from preprocess_card_summary import preprocess_card_summary_dynamic, create_highlighted_card_summary_dynamic
    from matching_helpers import identify_card_type, TransactionMatcher, filter_by_amount_range
    from highlighting_functions import create_highlighted_bank_statement, extract_matched_info_from_results
    # ADD THIS IMPORT
    from exclusive_discrepancy import calculate_total_discrepancies_by_card_type_exclusive
    
    print("=== Credit Card Transaction Matching ===\n")
    
    # Load data
    bank_statement = preprocess_bank_statement(bank_statement_path)
    card_summary, structure_info = preprocess_card_summary_dynamic(card_summary_path)
    
    # Prepare data
    bank_statement['Bank_Row_Number'] = range(2, len(bank_statement) + 2)
    bank_statement['Card_Type'] = bank_statement['Description'].apply(identify_card_type)
    
    # Create matcher with card-specific filters
    matcher = TransactionMatcher()
    matcher.add_filter('amount_range', filter_by_amount_range, card_types=['Amex'])
    
    # Run matching
    results = matcher.match_transactions(card_summary, bank_statement, 
                                        forward_days=3, verbose=verbose)
    
    # Extract info and generate reports
    matched_bank_rows, matched_dates_and_types, differences_by_row, differences_by_date_type, unmatched_info = extract_matched_info_from_results(results)
    
    # ADD THIS: Calculate discrepancies by card type
    discrepancies_by_type, first_matched_date = calculate_total_discrepancies_by_card_type_exclusive(
        results, bank_statement, matched_bank_rows
    )
    
    # Print discrepancy summary (optional)
    if verbose:
        print("\n=== CARD TYPE DISCREPANCIES ===")
        if first_matched_date:
            print(f"(Calculated from {first_matched_date.strftime('%Y-%m-%d')} onwards)")
        for card_type, disc in sorted(discrepancies_by_type.items()):
            if abs(disc) > 0.01:
                if disc > 0:
                    print(f"{card_type}: +${disc:,.2f} (bank has more)")
                else:
                    print(f"{card_type}: -${abs(disc):,.2f} (bank has less)")
    
    # Create highlighted files
    create_highlighted_bank_statement(
        bank_statement_path=bank_statement_path,
        matched_bank_rows=matched_bank_rows,
        output_path=f'{output_dir}/bank_statement_cards_highlighted.xlsx',
        differences_by_row=differences_by_row
    )
    
    # MODIFY THIS CALL to include differences_by_card_type
    create_highlighted_card_summary_dynamic(
        card_summary_path=card_summary_path,
        matched_dates_and_types=matched_dates_and_types,
        output_path=f'{output_dir}/card_summary_highlighted.xlsx',
        differences_info=differences_by_date_type,
        unmatched_info=unmatched_info,
        differences_by_card_type=discrepancies_by_type  # ADD THIS PARAMETER
    )
    
    print(f"✓ Card matching complete. Files saved to {output_dir}/")
    
    # MODIFY RETURN to include discrepancies for use in combined analysis
    return results, discrepancies_by_type, first_matched_date

def run_deposit_matching(deposit_slip_path: str, bank_statement_path: str, 
                        output_dir: str = '.', verbose: bool = False):
    """
    Run the deposit slip matching process.
    """
    from deposit_matching import process_deposit_slip
    
    results = process_deposit_slip(
        deposit_slip_path=deposit_slip_path,
        bank_statement_path=bank_statement_path,
        output_dir=output_dir,
        verbose=verbose
    )
    
    return results

def run_combined_analysis(card_summary_path: str, deposit_slip_path: str, 
                         bank_statement_path: str, output_dir: str = '.', 
                         verbose: bool = False):
    """
    Run both card and deposit matching, then create a combined analysis.
    """
    from preprocess_bank_statement import preprocess_bank_statement
    
    print("="*60)
    print("COMPREHENSIVE BANK RECONCILIATION SYSTEM")
    print("="*60)
    
    # Load bank statement once
    bank_statement = preprocess_bank_statement(bank_statement_path)
    total_bank_amount = bank_statement['Amount'].sum()
    
    print(f"\nBank Statement Summary:")
    print(f"  Total transactions: {len(bank_statement)}")
    print(f"  Date range: {bank_statement['Date'].min().strftime('%Y-%m-%d')} to {bank_statement['Date'].max().strftime('%Y-%m-%d')}")
    print(f"  Total amount: ${total_bank_amount:,.2f}")
    
    print("\n" + "-"*60)
    
    # Run card matching - NOW RECEIVES THREE VALUES
    card_results, card_discrepancies, first_matched_date = run_card_matching(
        card_summary_path, bank_statement_path, output_dir, verbose
    )
    
    print("\n" + "-"*60)
    
    # Run deposit matching
    deposit_results = run_deposit_matching(deposit_slip_path, bank_statement_path, 
                                          output_dir, verbose)
    
    # Combined summary
    print("\n" + "="*60)
    print("COMBINED RECONCILIATION SUMMARY")
    print("="*60)
    
    # Extract matched rows from both processes
    card_matched_rows = set()
    for date, date_results in card_results.items():
        # Skip metadata keys that start with underscore
        if isinstance(date, str) and date.startswith('_'):
            continue
        # Check if the expected key exists
        if 'matches_by_card_type' in date_results:
            for match_info in date_results['matches_by_card_type'].values():
                card_matched_rows.update(match_info['bank_rows'])
    
    deposit_matched_rows = deposit_results.get('_matched_bank_rows', set())
    
    # Calculate overlap (if any)
    overlap_rows = card_matched_rows & deposit_matched_rows
    all_matched_rows = card_matched_rows | deposit_matched_rows
    
    # Reload bank statement to check coverage
    bank_statement['Bank_Row_Number'] = range(2, len(bank_statement) + 2)
    matched_amount = bank_statement[bank_statement['Bank_Row_Number'].isin(all_matched_rows)]['Amount'].sum()
    unmatched_amount = bank_statement[~bank_statement['Bank_Row_Number'].isin(all_matched_rows)]['Amount'].sum()
    
    print(f"\nMatching Coverage:")
    print(f"  Card matching: {len(card_matched_rows)} transactions")
    print(f"  Deposit matching: {len(deposit_matched_rows)} transactions")
    if overlap_rows:
        print(f"  ⚠ Overlap: {len(overlap_rows)} transactions matched by both")
        print(f"    Bank rows: {sorted(list(overlap_rows)[:10])}")
    print(f"  Total unique matched: {len(all_matched_rows)} transactions")
    print(f"  Total unmatched: {len(bank_statement) - len(all_matched_rows)} transactions")
    
    print(f"\nFinancial Coverage:")
    print(f"  Matched amount: ${matched_amount:,.2f} ({matched_amount/total_bank_amount*100:.1f}%)")
    print(f"  Unmatched amount: ${unmatched_amount:,.2f} ({unmatched_amount/total_bank_amount*100:.1f}%)")
    
    # NEW SECTION: Display card discrepancies
    if card_discrepancies:
        print("\n" + "-"*60)
        print("CARD TYPE DISCREPANCIES")
        print("-"*60)
        if first_matched_date:
            print(f"(Calculated from {first_matched_date.strftime('%Y-%m-%d')} onwards)\n")
        
        total_card_discrepancy = 0
        for card_type, disc in sorted(card_discrepancies.items()):
            if abs(disc) > 0.01:  # Only show non-zero discrepancies
                if disc > 0:
                    print(f"  {card_type}: +${disc:,.2f} (bank has more than expected)")
                else:
                    print(f"  {card_type}: -${abs(disc):,.2f} (bank has less than expected)")
                total_card_discrepancy += disc
        
        print(f"\n  Total net discrepancy: ${total_card_discrepancy:,.2f}")
        if total_card_discrepancy > 0:
            print("  (Positive = bank statement total exceeds card summary expectations)")
        elif total_card_discrepancy < 0:
            print("  (Negative = card summary expects more than bank statement shows)")
    
    # Create combined highlighted bank statement
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    import shutil
    
    combined_output = f'{output_dir}/bank_statement_combined_highlighted.xlsx'
    
    # Write bank statement to Excel
    with pd.ExcelWriter(combined_output, engine='openpyxl') as writer:
        bank_statement.to_excel(writer, sheet_name='Bank Statement', index=False)
    
    # Load and apply highlighting
    workbook = load_workbook(combined_output)
    worksheet = workbook['Bank Statement']
    
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Card matches
    blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')   # Deposit matches
    purple_fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid') # Both
    
    for row_num in range(2, len(bank_statement) + 2):
        if row_num in overlap_rows:
            # Purple for rows matched by both
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_num, column=col).fill = purple_fill
        elif row_num in card_matched_rows:
            # Green for card matches
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_num, column=col).fill = green_fill
        elif row_num in deposit_matched_rows:
            # Blue for deposit matches
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_num, column=col).fill = blue_fill
    
    workbook.save(combined_output)
    workbook.close()
    
    print(f"\n✓ Combined highlighted bank statement: {combined_output}")
    print("  Color legend:")
    print("    Green = Matched by card reconciliation")
    print("    Blue = Matched by deposit reconciliation")
    print("    Purple = Matched by both (verify if this is correct)")
    print("    No color = Unmatched")
    
    print("\n" + "="*60)
    print("ALL PROCESSING COMPLETE")
    print("="*60)
    
    # Generate comprehensive summary report
    print("\nGenerated files:")
    print(f"  1. {output_dir}/bank_statement_cards_highlighted.xlsx")
    print(f"  2. {output_dir}/card_summary_highlighted.xlsx (with discrepancy notes)")
    print(f"  3. {output_dir}/deposit_slip_highlighted.xlsx")
    print(f"  4. {output_dir}/bank_statement_combined_highlighted.xlsx")
    
    if first_matched_date:
        print(f"\nNote: Card discrepancies calculated from {first_matched_date.strftime('%Y-%m-%d')} onwards")
        print("(Earlier bank transactions excluded as they belong to previous month)")

def main():
    parser = argparse.ArgumentParser(description='Bank Reconciliation System')
    parser.add_argument('--mode', choices=['cards', 'deposits', 'both'], 
                       default='both', help='What to reconcile')
    parser.add_argument('--bank', required=True, help='Bank statement CSV path')
    parser.add_argument('--cards', help='Card summary Excel path')
    parser.add_argument('--deposits', help='Deposit slip Excel path')
    parser.add_argument('--output', default='.', help='Output directory')
    parser.add_argument('--verbose', action='store_true', help='Verbose output')
    
    args = parser.parse_args()
    
    if args.mode == 'cards':
        if not args.cards:
            print("Error: --cards path required for cards mode")
            return
        run_card_matching(args.cards, args.bank, args.output, args.verbose)
        
    elif args.mode == 'deposits':
        if not args.deposits:
            print("Error: --deposits path required for deposits mode")
            return
        run_deposit_matching(args.deposits, args.bank, args.output, args.verbose)
        
    else:  # both
        if not args.cards or not args.deposits:
            print("Error: Both --cards and --deposits paths required for combined mode")
            return
        run_combined_analysis(args.cards, args.deposits, args.bank, 
                            args.output, args.verbose)

if __name__ == "__main__":
    # For testing, run with hardcoded paths
    import os
    
    if len(os.sys.argv) == 1:
        # No command line args - run test
        print("Running with test configuration...")
        run_combined_analysis(
            card_summary_path='data/XYZ Storage Laird - CreditCardSummary - 06-01-2025 - 06-30-2025.xlsx',
            deposit_slip_path='data/XYZ Storage Laird - MonthlyDepositSlip - 06-01-2025 - 06-30-2025 (2).xlsx',
            bank_statement_path='data/june 2025 bank statement.CSV',
            output_dir='.',
            verbose=True
        )
    else:
        # Command line args provided
        main()