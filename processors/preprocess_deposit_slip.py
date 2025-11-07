import pandas as pd
import numpy as np

def detect_deposit_slip_structure(filepath):
    """
    Dynamically detect the structure of a deposit slip file.
    Similar to card summary but adapted for deposit slip format.
    """
    # Read all rows first to analyze structure
    df_raw = pd.read_excel(filepath, header=None)
    
    # Find the header row by looking for 'Date' in first column
    header_row = None
    for idx, row in df_raw.iterrows():
        if pd.notna(row[0]) and 'Date' in str(row[0]):
            header_row = idx
            break
    
    if header_row is None:
        raise ValueError("Could not find header row with 'Date'")
    
    # Find the first data row (skip facility name rows)
    data_start_row = None
    for idx in range(header_row + 1, len(df_raw)):
        try:
            # Check if this looks like a date (not a facility name)
            cell_value = str(df_raw.iloc[idx, 0]).strip()
            if 'Facility Name:' in cell_value or 'facility' in cell_value.lower():
                continue
            pd.to_datetime(cell_value)
            data_start_row = idx
            break
        except:
            continue
    
    # Find the total row (look for the LAST "Total" in the first column)
    total_row = None
    for idx in range(len(df_raw) - 1, data_start_row - 1, -1):  # Search backwards
        if pd.notna(df_raw.iloc[idx, 0]) and str(df_raw.iloc[idx, 0]).strip() == 'Total':
            total_row = idx
            break
    
    # Find any facility name rows and intermediate total rows
    facility_rows = []
    intermediate_total_rows = []
    for idx in range(data_start_row, len(df_raw)):
        if pd.notna(df_raw.iloc[idx, 0]):
            cell_value = str(df_raw.iloc[idx, 0]).strip()
            if 'Facility Name:' in cell_value:
                facility_rows.append(idx)
            elif cell_value == 'Total' and idx != total_row:  # Skip the final total row
                intermediate_total_rows.append(idx)
    
    # Build skip rows list
    skip_rows = []
    skip_rows.extend(range(header_row))
    for idx in range(header_row + 1, data_start_row):
        skip_rows.append(idx)
    skip_rows.extend(facility_rows)  # Add facility name rows
    skip_rows.extend(intermediate_total_rows)  # Add intermediate total rows
    if total_row is not None:
        skip_rows.append(total_row)
    
    print(f"Deposit slip structure detected:")
    print(f"  Header row: {header_row}")
    print(f"  Data starts: {data_start_row}")
    print(f"  Total row: {total_row}")
    print(f"  Skip rows: {skip_rows}")
    
    return skip_rows, header_row, data_start_row, total_row

def preprocess_deposit_slip_dynamic(filepath):
    """
    Dynamically load and preprocess deposit slip Excel file.
    Handles Cash and Check columns specifically.
    """
    # Detect structure
    skip_rows, header_row, data_start_row, total_row = detect_deposit_slip_structure(filepath)
    
    # Load the Excel file
    deposit_slip = pd.read_excel(filepath, skiprows=skip_rows)
    
    # Convert date to datetime
    deposit_slip['Date'] = pd.to_datetime(deposit_slip['Date'])
    
    # Clean column names
    deposit_slip.columns = deposit_slip.columns.str.strip()
    
    # Identify Cash and Check columns (adjust these names based on your actual file)
    deposit_columns = []
    for col in deposit_slip.columns:
        if 'Cash' in col or 'Check' in col:
            deposit_columns.append(col)
    
    # Convert numeric columns
    for col in deposit_columns:
        if col in deposit_slip.columns:
            deposit_slip[col] = pd.to_numeric(deposit_slip[col], errors='coerce').fillna(0)
    
    # Add a Total column if not present
    if 'Total' not in deposit_slip.columns and deposit_columns:
        deposit_slip['Total'] = deposit_slip[deposit_columns].sum(axis=1)
    
    return deposit_slip, {
        'skip_rows': skip_rows,
        'header_row': header_row,
        'data_start_row': data_start_row,
        'total_row': total_row,
        'deposit_columns': deposit_columns
    }

def create_highlighted_deposit_slip(deposit_slip_path: str, 
                                  matched_dates_and_types: dict,
                                  output_path: str = 'deposit_slip_highlighted.xlsx',
                                  unmatched_info: dict = None,
                                  gc_allocation: dict = None,
                                  deposit_discrepancies: dict = None):
    """
    Create highlighted deposit slip with special handling for GC allocations.
    
    Args:
        deposit_slip_path: Path to original deposit slip
        matched_dates_and_types: Dict of matched dates and deposit types
        output_path: Output path for highlighted file
        unmatched_info: Information about unmatched entries
        gc_allocation: Dict showing how GC transactions were allocated between Cash/Check
    """
    import shutil
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment
    
    # Get the structure info
    deposit_slip_df, structure_info = preprocess_deposit_slip_dynamic(deposit_slip_path)
    
    # Copy the original Excel file
    shutil.copy2(deposit_slip_path, output_path)
    
    # Load the workbook
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    
    # Define highlight colors (text colors instead of background)
    green_font = Font(color='006400')  # Dark green for matched
    red_font = Font(color='DC143C')    # Crimson red for unmatched
    yellow_font = Font(color='B8860B') # Dark goldenrod for special cases
    # Create mapping of dates to Excel rows
    excel_row_mapping = {}
    excel_row = structure_info['data_start_row'] + 1
    
    for idx, row in deposit_slip_df.iterrows():
        while (excel_row - 1) in structure_info['skip_rows']:
            excel_row += 1
        excel_row_mapping[row['Date']] = excel_row
        excel_row += 1
    
    # Find column indices for deposit types
    header_row = structure_info['header_row'] + 1
    column_mapping = {}
    
    for col_idx in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col_idx).value
        if cell_value:
            column_mapping[str(cell_value).strip()] = col_idx
    
    # Process all data rows
    matched_cells_count = 0
    unmatched_cells_count = 0
    gc_1416_cells_count = 0
    
    for _, row in deposit_slip_df.iterrows():
        date = row['Date']
        excel_row = excel_row_mapping[date]
        
        # Process Cash and Check columns
        for deposit_type in structure_info['deposit_columns']:
            if deposit_type in column_mapping and deposit_type in row.index:
                expected_amount = row[deposit_type]
                if pd.notna(expected_amount) and expected_amount != 0:
                    col_idx = column_mapping[deposit_type]
                    cell = worksheet.cell(row=excel_row, column=col_idx)
                    
                    # Check if matched
                    if matched_dates_and_types and date in matched_dates_and_types:
                        if deposit_type in matched_dates_and_types[date]:
                            cell.font = green_font
                            matched_cells_count += 1
                            
                            # Add comment if this was from GC allocation
                            if gc_allocation and (date, deposit_type) in gc_allocation:
                                alloc_info = gc_allocation[(date, deposit_type)]
                                comment_text = f"Matched from GC transactions:\n"
                                comment_text += f"Allocated ${alloc_info['amount']:,.2f} to {deposit_type}\n"
                                comment_text += f"Bank rows: {', '.join(map(str, alloc_info['bank_rows'][:5]))}"
                                if len(alloc_info['bank_rows']) > 5:
                                    comment_text += f"... and {len(alloc_info['bank_rows'])-5} more"
                                cell.comment = Comment(comment_text, "GC Allocation")
                                gc_1416_cells_count += 1
                        else:
                            # Check if partially matched (e.g., Cash matched but Check didn't)
                            if any(dt in matched_dates_and_types[date] for dt in structure_info['deposit_columns']):
                                cell.font = yellow_font  # Partial match for the date

                    # NEW: Check for best match (yellow highlighting)
                    elif unmatched_info and (date, deposit_type) in unmatched_info:
                        info = unmatched_info[(date, deposit_type)]
                        
                        # Check if this has a best_match stored
                        if 'best_match' in info and info['best_match'] and info['best_match']['total'] > 0:
                            # Yellow for approximate matches
                            cell.font = yellow_font
                            best_match = info['best_match']
                            
                            # Add detailed comment
                            comment_text = f"Expected: ${expected_amount:,.2f}\n"
                            comment_text += f"Best match found: ${best_match['total']:.2f}\n"
                            comment_text += f"Difference: ${best_match['difference']:.2f}\n"
                            comment_text += f"Using {best_match['combo_size']} GC transaction(s)\n"
                            comment_text += f"Bank rows: {', '.join(map(str, best_match['bank_rows'][:5]))}"
                            if len(best_match['bank_rows']) > 5:
                                comment_text += f"... and {len(best_match['bank_rows'])-5} more"
                            
                            cell.comment = Comment(comment_text, "Best Match (Not Exact)")
                        else:
                            # Red for no matches at all
                            cell.font = red_font
                            unmatched_cells_count += 1
                            
                            comment_text = f"Expected: ${expected_amount:,.2f}\n"
                            comment_text += f"No GC transactions found\n"
                            comment_text += info.get('reason', 'No matches in date range')
                            
                            cell.comment = Comment(comment_text, "Unmatched")
                    
                    elif unmatched_info and (date, deposit_type) in unmatched_info:
                        # Unmatched cell
                        cell.font = red_font
                        unmatched_cells_count += 1
                        
                        info = unmatched_info[(date, deposit_type)]
                        comment_text = f"Expected: ${expected_amount:,.2f}\n"
                        comment_text += f"Found GC: ${info.get('gc_found', 0):,.2f}\n"
                        comment_text += f"Difference: ${info.get('gc_found', 0) - expected_amount:,.2f}\n"
                        comment_text += "\nNote: GC transactions can be either Cash or Check"
                        
                        cell.comment = Comment(comment_text, "Matching System")
    
    # ADD NEW DISCREPANCY ROW below the total row
    if structure_info['total_row'] is not None and deposit_discrepancies:
        from openpyxl.styles import Font
        discrepancy_row = structure_info['total_row'] + 2  # +2 to be below total row
        
        # Add label in first column
        label_cell = worksheet.cell(row=discrepancy_row, column=1)
        label_cell.value = "Net Discrepancy"
        label_cell.font = Font(bold=True)
        
        # Add discrepancy values for each deposit type
        for deposit_type, col_idx in column_mapping.items():
            if deposit_type in deposit_discrepancies and deposit_type in ['Cash', 'Check']:
                diff = deposit_discrepancies[deposit_type]
                cell = worksheet.cell(row=discrepancy_row, column=col_idx)
                
                # Format positive/negative with appropriate colors
                if abs(diff) > 0.01:
                    cell.value = diff
                    cell.number_format = '$#,##0.00'
                    
                    # Color code: red for negative (bank has less), blue for positive (bank has more)
                    if diff > 0:
                        cell.font = Font(color='0000FF', bold=True)  # Blue for positive
                    else:
                        cell.font = Font(color='FF0000', bold=True)  # Red for negative
                else:
                    cell.value = 0
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(color='008000', bold=True)  # Green for zero
    
    # Save
    workbook.save(output_path)
    workbook.close()
    
    print(f"✓ Created highlighted deposit slip: {output_path}")
    print(f"  - Detected {len(deposit_slip_df)} data rows")
    print(f"  - Matched cells (green): {matched_cells_count}")
    print(f"  - GC allocations: {gc_1416_cells_count}")
    print(f"  - Unmatched cells (red): {unmatched_cells_count}")
    if deposit_discrepancies:
        non_zero_diffs = sum(1 for d in deposit_discrepancies.values() if abs(d) > 0.01)
        if non_zero_diffs > 0:
            print(f"  - Discrepancy row added with {non_zero_diffs} non-zero values")

if __name__ == "__main__":
    # Test the preprocessor
    test_file = 'XYZ Storage Laird - MonthlyDepositSlip - 06-01-2025 - 06-30-2025.xlsx'
    
    try:
        print(f"Testing {test_file}:")
        df, info = preprocess_deposit_slip_dynamic(test_file)
        print(f"Successfully loaded {len(df)} days")
        print(f"Date range: {df['Date'].min()} to {df['Date'].max()}")
        print(f"Deposit columns found: {info['deposit_columns']}")
        print("\nFirst few rows:")
        print(df.head())
    except Exception as e:
        print(f"Error: {e}")