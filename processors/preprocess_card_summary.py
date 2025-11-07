import pandas as pd
import numpy as np

def detect_card_summary_structure(filepath='card summary june.xlsx'):
    """
    Dynamically detect the structure of a card summary file.
    Returns the data rows and rows to skip.
    """
    # Read all rows first to analyze structure
    df_raw = pd.read_excel(filepath, header=None)
    
    # Find the header row by looking for 'Date' in first column
    header_row = None
    for idx, row in df_raw.iterrows():
        if pd.notna(row[0]) and 'Date' in str(row[0]):
            header_row = idx
            break
    
    if header_row is None:
        raise ValueError("Could not find header row with 'Date'")
    
    # Find the first data row (first valid date after header)
    data_start_row = None
    for idx in range(header_row + 1, len(df_raw)):
        try:
            # Try to parse as date
            pd.to_datetime(df_raw.iloc[idx, 0])
            data_start_row = idx
            break
        except:
            continue
    
    # Find the total row by looking for 'Total' in first column
    total_row = None
    for idx in range(data_start_row, len(df_raw)):
        if pd.notna(df_raw.iloc[idx, 0]) and 'Total' in str(df_raw.iloc[idx, 0]):
            total_row = idx
            break
    
    # Find any empty rows between header and data start
    skip_rows = []
    
    # Add rows before header
    skip_rows.extend(range(header_row))
    
    # Add empty rows between header and data
    for idx in range(header_row + 1, data_start_row):
        skip_rows.append(idx)
    
    # Add total row if found
    if total_row is not None:
        skip_rows.append(total_row)
    
    print(f"Structure detected:")
    print(f"  Header row: {header_row}")
    print(f"  Data starts: {data_start_row}")
    print(f"  Total row: {total_row}")
    print(f"  Skip rows: {skip_rows}")
    
    return skip_rows, header_row, data_start_row, total_row

def preprocess_card_summary_dynamic(filepath='card summary june.xlsx'):
    """
    Dynamically load and preprocess any card summary Excel file.
    """
    # Detect structure
    skip_rows, header_row, data_start_row, total_row = detect_card_summary_structure(filepath)
    
    # Load the Excel file with detected skip rows
    card_summary = pd.read_excel(filepath, skiprows=skip_rows)
    
    # Convert date to datetime
    card_summary['Date'] = pd.to_datetime(card_summary['Date'])
    
    # Clean column names
    card_summary.columns = card_summary.columns.str.strip()
    
    # Convert numeric columns
    numeric_columns = [col for col in card_summary.columns 
                      if col not in ['Date'] and not col.startswith('Unnamed')]
    
    for col in numeric_columns:
        card_summary[col] = pd.to_numeric(card_summary[col], errors='coerce').fillna(0)
    
    # Return both the data and the structure info
    return card_summary, {
        'skip_rows': skip_rows,
        'header_row': header_row,
        'data_start_row': data_start_row,
        'total_row': total_row
    }

# Updated highlighting function that uses dynamic structure
def create_highlighted_card_summary_dynamic(card_summary_path: str, matched_dates_and_types: dict,
                                          output_path: str = 'card_summary_highlighted.xlsx',
                                          differences_info: dict = None,
                                          unmatched_info: dict = None, 
                                          unmatched_bank_by_type: dict = None,
                                          differences_by_card_type: dict = None):
    """
    Create highlighted card summary with dynamic structure detection.
    Now with simplified total row comments showing just the sum of differences.
    """
    import shutil
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment
    import openpyxl.styles
    
    # Get the structure info
    card_summary_df, structure_info = preprocess_card_summary_dynamic(card_summary_path)
    
    # Copy the original Excel file
    shutil.copy2(card_summary_path, output_path)
    
    # Load the workbook
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    
    # Define highlight colors (text colors instead of background)
    green_font = Font(color='006400')  # Dark green for matched
    red_font = Font(color='DC143C')    # Crimson red for unmatched
    
    # Create mapping of dates to Excel rows
    excel_row_mapping = {}
    excel_row = structure_info['data_start_row'] + 1  # +1 for 1-based Excel rows
    
    for idx, row in card_summary_df.iterrows():
        # Skip rows that were skipped in original
        while (excel_row - 1) in structure_info['skip_rows']:
            excel_row += 1
        excel_row_mapping[row['Date']] = excel_row
        excel_row += 1
    
    # Find column indices for card types
    header_row = structure_info['header_row'] + 1  # +1 for 1-based Excel rows
    column_mapping = {}
    
    for col_idx in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col_idx).value
        if cell_value:
            column_mapping[str(cell_value).strip()] = col_idx
    
    # Process all data rows (not including total)
    matched_cells_count = 0
    unmatched_cells_count = 0
    
    for _, row in card_summary_df.iterrows():
        date = row['Date']
        excel_row = excel_row_mapping[date]
        
        # Process each card type
        for card_type in column_mapping.keys():
            if card_type in ['Date', 'Total', 'Visa & MC'] or card_type.startswith('Unnamed'):
                continue
                
            if card_type in column_mapping and card_type in row.index:
                expected_amount = row[card_type]
                if pd.notna(expected_amount) and expected_amount != 0:
                    col_idx = column_mapping[card_type]
                    cell = worksheet.cell(row=excel_row, column=col_idx)
                    
                    # Check if matched
                    if matched_dates_and_types and date in matched_dates_and_types and card_type in matched_dates_and_types[date]:
                        cell.font = green_font
                        matched_cells_count += 1
                    elif unmatched_info and (date, card_type) in unmatched_info:
                        cell.font = red_font
                        unmatched_cells_count += 1
                        
                        # Add comment
                        info = unmatched_info[(date, card_type)]
                        comment_text = f"Expected: ${expected_amount:,.2f}\n"
                        comment_text += f"Found: ${info.get('total_found', 0):,.2f}\n"
                        comment_text += f"Difference: ${info.get('total_found', 0) - expected_amount:,.2f}\n"
                        comment_text += f"Unmatched transactions available: {info.get('found_transactions', 0)}\n"
                        comment_text += "\nNote: 'Found' amount excludes transactions already matched elsewhere"
                        
                        cell.comment = Comment(comment_text, "Matching System")
    
    # Handle total row with simplified differences if provided
    if structure_info['total_row'] is not None and differences_by_card_type:
        total_excel_row = structure_info['total_row'] + 1  # +1 for Excel
        
        for card_type, col_idx in column_mapping.items():
            if card_type in differences_by_card_type and card_type not in ['Date', 'Total', 'Visa & MC']:
                diff = differences_by_card_type[card_type]
                # Only add comment if there's a non-zero difference
                if abs(diff) > 0.01:
                    cell = worksheet.cell(row=total_excel_row, column=col_idx)
                    
                    comment_text = f"Net Discrepancy:\n${diff:,.2f}"
                    if diff > 0:
                        comment_text += "\n\n(Bank has ${abs(diff):,.2f} MORE than expected)"
                        comment_text += "\nPossible causes: duplicate transactions,\nunexpected charges, or misclassified items"
                    else:
                        comment_text += "\n\n(Bank has ${abs(diff):,.2f} LESS than expected)"
                        comment_text += "\nPossible causes: missing transactions,\nunprocessed charges, or timing differences"
                    
                    cell.comment = Comment(comment_text, "Difference Summary")
    
    # ADD NEW DISCREPANCY ROW below the total row
    if structure_info['total_row'] is not None and differences_by_card_type:
        discrepancy_row = structure_info['total_row'] + 2  # +2 to be below total row
        
        # Add label in first column
        label_cell = worksheet.cell(row=discrepancy_row, column=1)
        label_cell.value = "Net Discrepancy"
        label_cell.font = Font(bold=True)
        
        # Add discrepancy values for each card type
        for card_type, col_idx in column_mapping.items():
            if card_type in differences_by_card_type and card_type not in ['Date', 'Total', 'Visa & MC']:
                diff = differences_by_card_type[card_type]
                cell = worksheet.cell(row=discrepancy_row, column=col_idx)
                
                # Format positive/negative with appropriate colors
                if abs(diff) > 0.01:
                    cell.value = diff
                    cell.number_format = '$#,##0.00'
                    
                    # Color code: red for negative (bank has less), blue for positive (bank has more)
                    if diff > 0:
                        cell.font = Font(color='0000FF', bold=True)  # Blue for positive
                    else:
                        cell.font = Font(color='FF0000', bold=True)  # Red for negative
                else:
                    cell.value = 0
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(color='008000', bold=True)  # Green for zero
    
    # Save
    workbook.save(output_path)
    workbook.close()
    
    print(f"✓ Created highlighted card summary: {output_path}")
    print(f"  - Detected {len(card_summary_df)} data rows")
    print(f"  - Matched cells (green): {matched_cells_count}")
    print(f"  - Unmatched cells with comments (red): {unmatched_cells_count}")
    if differences_by_card_type:
        non_zero_diffs = sum(1 for d in differences_by_card_type.values() if abs(d) > 0.01)
        if non_zero_diffs > 0:
            print(f"  - Total row differences shown for {non_zero_diffs} card types")

if __name__ == "__main__":
    # Test with different files
    test_files = [
        'card summary june.xlsx',  # 30 days
        # 'card summary july.xlsx',  # 31 days
        # 'card summary feb.xlsx',   # 28/29 days
    ]
    
    for file in test_files:
        try:
            print(f"\nTesting {file}:")
            df, info = preprocess_card_summary_dynamic(file)
            print(f"Successfully loaded {len(df)} days")
            print(f"Date range: {df['Date'].min()} to {df['Date'].max()}")
        except Exception as e:
            print(f"Error: {e}")