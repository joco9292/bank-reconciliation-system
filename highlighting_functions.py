import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import shutil

def create_highlighted_bank_statement(bank_statement_path: str, matched_bank_rows: set, 
                                    output_path: str = 'bank_statement_highlighted.xlsx',
                                    differences_by_row: dict = None,
                                    transaction_details: dict = None,
                                    match_type_info: dict = None,
                                    unmatched_transactions: dict = None,
                                    gc_transactions: dict = None):
    """
    Create a highlighted copy of the bank statement with matched rows highlighted and informative comments.
    
    Args:
        bank_statement_path (str): Path to original bank statement CSV
        matched_bank_rows (set): Set of row numbers that were matched (1-based Excel rows)
        output_path (str): Output path for highlighted Excel file
        differences_by_row (dict): Optional dict mapping bank rows to difference amounts
        transaction_details (dict): Optional dict mapping bank rows to transaction details for comments
        match_type_info (dict): Optional dict mapping bank rows to match type information
        unmatched_transactions (dict): Optional dict mapping bank rows to unmatched transaction info
        gc_transactions (dict): Optional dict mapping bank rows to GC transaction info
    """
    # Read the original CSV file
    original_df = pd.read_csv(bank_statement_path)
    
    # Write to Excel with highlighting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        original_df.to_excel(writer, sheet_name='Bank Statement', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Bank Statement']
        
        # Define highlight colors
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Matched
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')    # Unmatched
        blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')   # GC transactions
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid') # Other special cases
        
        # Process all rows in the bank statement
        for row_idx in range(len(original_df)):
            excel_row = row_idx + 2  # +2 because Excel is 1-based and has header
            worksheet_row = excel_row
            
            if 2 <= worksheet_row <= len(original_df) + 1:
                # Determine highlight color and comment based on transaction type
                fill_color = None
                comment_text = None
                
                if excel_row in matched_bank_rows:
                    # Matched transaction
                    fill_color = green_fill
                    if transaction_details and excel_row in transaction_details:
                        comment_text = _create_transaction_comment(
                            excel_row, transaction_details[excel_row], 
                            differences_by_row.get(excel_row) if differences_by_row else None,
                            match_type_info.get(excel_row) if match_type_info else None
                        )
                elif unmatched_transactions and excel_row in unmatched_transactions:
                    # Unmatched transaction
                    fill_color = red_fill
                    comment_text = _create_unmatched_transaction_comment(
                        excel_row, unmatched_transactions[excel_row]
                    )
                elif gc_transactions and excel_row in gc_transactions:
                    # GC transaction
                    fill_color = blue_fill
                    comment_text = _create_gc_transaction_comment(
                        excel_row, gc_transactions[excel_row]
                    )
                
                # Apply highlighting and comments
                if fill_color:
                    for col in range(1, len(original_df.columns) + 1):
                        cell = worksheet.cell(row=worksheet_row, column=col)
                        cell.fill = fill_color
                        
                        # Add comment to the first cell
                        if col == 1 and comment_text:
                            cell.comment = Comment(comment_text, "Bank Reconciliation System")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✓ Created highlighted bank statement: {output_path}")
    print(f"  - Total rows: {len(original_df)}")
    print(f"  - Matched rows: {len(matched_bank_rows)}")
    if transaction_details:
        print(f"  - Rows with comments: {len(transaction_details)}")

def _create_transaction_comment(bank_row: int, transaction_info: dict, 
                              difference: float = None, match_type: str = None) -> str:
    """
    Create an informative comment for a matched transaction showing what it matched from.
    
    Args:
        bank_row (int): Bank statement row number
        transaction_info (dict): Transaction details from matching results
        difference (float): Optional difference amount
        match_type (str): Optional match type information
        
    Returns:
        str: Formatted comment text
    """
    comment_lines = []
    
    # Show what this bank transaction matched from
    if 'Expected_Amount' in transaction_info:
        comment_lines.append(f"Matched from Card Summary:")
        comment_lines.append(f"  Date: {transaction_info.get('Date', 'Unknown')}")
        comment_lines.append(f"  Card Type: {transaction_info.get('Card_Type', 'Unknown')}")
        comment_lines.append(f"  Expected: ${transaction_info['Expected_Amount']:,.2f}")
        comment_lines.append(f"  Actual: ${transaction_info.get('Actual_Total', 0):,.2f}")
    
    # Match information
    if match_type:
        comment_lines.append(f"Match Type: {match_type}")
    
    # Difference information
    if difference is not None and abs(difference) > 0.01:
        if difference > 0:
            comment_lines.append(f"Difference: +${difference:,.2f} (bank has more)")
        else:
            comment_lines.append(f"Difference: -${abs(difference):,.2f} (bank has less)")
    
    # Additional context
    comment_lines.append(f"Bank Row: {bank_row}")
    comment_lines.append("Status: Matched")
    
    return "\n".join(comment_lines)

def _create_combined_transaction_comment(bank_row: int, card_details: dict, deposit_details: dict,
                                       card_match_type: str = None, deposit_match_type: str = None) -> str:
    """
    Create an informative comment for a transaction matched by both card and deposit systems.
    
    Args:
        bank_row (int): Bank statement row number
        card_details (dict): Card transaction details
        deposit_details (dict): Deposit transaction details
        card_match_type (str): Card match type
        deposit_match_type (str): Deposit match type
        
    Returns:
        str: Formatted comment text
    """
    comment_lines = []
    
    # Show what this bank transaction matched from both sources
    comment_lines.append("Matched from BOTH sources:")
    
    if card_details:
        comment_lines.append("  Card Summary:")
        comment_lines.append(f"    Date: {card_details.get('Date', 'Unknown')}")
        comment_lines.append(f"    Card Type: {card_details.get('Card_Type', 'Unknown')}")
        comment_lines.append(f"    Expected: ${card_details.get('Expected_Amount', 0):,.2f}")
        comment_lines.append(f"    Actual: ${card_details.get('Actual_Total', 0):,.2f}")
    
    if deposit_details:
        comment_lines.append("  Deposit Slip:")
        comment_lines.append(f"    Date: {deposit_details.get('Date', 'Unknown')}")
        comment_lines.append(f"    Type: {deposit_details.get('Card_Type', 'Deposit')}")
        comment_lines.append(f"    Expected: ${deposit_details.get('Expected_Amount', 0):,.2f}")
        comment_lines.append(f"    Actual: ${deposit_details.get('Actual_Total', 0):,.2f}")
    
    # Match information
    comment_lines.append("Match Status: Matched by BOTH systems")
    if card_match_type:
        comment_lines.append(f"Card Match Type: {card_match_type}")
    if deposit_match_type:
        comment_lines.append(f"Deposit Match Type: {deposit_match_type}")
    
    # Additional context
    comment_lines.append(f"Bank Row: {bank_row}")
    comment_lines.append("Note: This bank transaction was matched from both card summary and deposit slip")
    
    return "\n".join(comment_lines)

def _create_unmatched_transaction_comment(bank_row: int, unmatched_info: dict) -> str:
    """
    Create an informative comment for an unmatched transaction showing what it tried to match from.
    
    Args:
        bank_row (int): Bank statement row number
        unmatched_info (dict): Unmatched transaction information
        
    Returns:
        str: Formatted comment text
    """
    comment_lines = []
    
    # Show what this bank transaction tried to match from
    comment_lines.append("Tried to match from Card Summary:")
    comment_lines.append(f"  Date: {unmatched_info.get('Date', 'Unknown')}")
    comment_lines.append(f"  Card Type: {unmatched_info.get('Card_Type', 'Unknown')}")
    comment_lines.append(f"  Expected: ${unmatched_info.get('expected', 0):,.2f}")
    comment_lines.append(f"  Found: ${unmatched_info.get('total_found', 0):,.2f}")
    
    # Unmatched information
    comment_lines.append("Status: UNMATCHED")
    
    if 'reason' in unmatched_info:
        comment_lines.append(f"Reason: {unmatched_info['reason']}")
    
    if 'filters_tried' in unmatched_info:
        comment_lines.append(f"Filters Tried: {', '.join(unmatched_info['filters_tried'])}")
    
    # Additional context
    comment_lines.append(f"Bank Row: {bank_row}")
    comment_lines.append("Note: This bank transaction could not be matched to any expected amount")
    
    return "\n".join(comment_lines)

def _create_gc_transaction_comment(bank_row: int, gc_info: dict) -> str:
    """
    Create an informative comment for a GC (Cash & Check) transaction showing what it matched from.
    
    Args:
        bank_row (int): Bank statement row number
        gc_info (dict): GC transaction information
        
    Returns:
        str: Formatted comment text
    """
    comment_lines = []
    
    # Show what this bank transaction matched from
    comment_lines.append("Matched from Deposit Slip:")
    comment_lines.append(f"  Date: {gc_info.get('Date', 'Unknown')}")
    comment_lines.append(f"  Transaction Type: GC (Cash & Check)")
    comment_lines.append(f"  Amount: ${gc_info.get('Amount', 0):,.2f}")
    
    # Allocation information
    if 'allocation' in gc_info:
        allocation = gc_info['allocation']
        if 'cash_amount' in allocation and allocation['cash_amount'] > 0:
            comment_lines.append(f"  Cash Allocation: ${allocation['cash_amount']:,.2f}")
        if 'check_amount' in allocation and allocation['check_amount'] > 0:
            comment_lines.append(f"  Check Allocation: ${allocation['check_amount']:,.2f}")
    
    if 'match_type' in gc_info:
        comment_lines.append(f"Match Type: {gc_info['match_type']}")
    
    if 'deposit_date' in gc_info:
        comment_lines.append(f"Deposit Date: {gc_info['deposit_date']}")
    
    # Additional context
    comment_lines.append(f"Bank Row: {bank_row}")
    comment_lines.append("Status: Matched")
    comment_lines.append("Note: This bank transaction was matched from the deposit slip")
    
    return "\n".join(comment_lines)

def create_combined_highlighted_bank_statement(bank_statement_path: str, 
                                             card_matched_rows: set = None,
                                             deposit_matched_rows: set = None,
                                             card_attempted_rows: set = None,
                                             deposit_attempted_rows: set = None,
                                             output_path: str = 'bank_statement_combined_highlighted.xlsx',
                                             card_transaction_details: dict = None,
                                             deposit_transaction_details: dict = None,
                                             card_match_type_info: dict = None,
                                             deposit_match_type_info: dict = None,
                                             card_unmatched_transactions: dict = None,
                                             deposit_unmatched_transactions: dict = None,
                                             gc_transactions: dict = None):
    """
    Create a combined highlighted copy of the original bank statement showing results from both 
    credit card and deposit matching.
    
    Args:
        bank_statement_path (str): Path to original bank statement CSV
        card_matched_rows (set): Set of row numbers matched by card matching (1-based Excel rows)
        deposit_matched_rows (set): Set of row numbers matched by deposit matching (1-based Excel rows)
        card_attempted_rows (set): Set of row numbers attempted by card matching (1-based Excel rows)
        deposit_attempted_rows (set): Set of row numbers attempted by deposit matching (1-based Excel rows)
        output_path (str): Output path for highlighted Excel file
    """
    # Initialize sets if None
    card_matched_rows = card_matched_rows or set()
    deposit_matched_rows = deposit_matched_rows or set()
    card_attempted_rows = card_attempted_rows or set()
    deposit_attempted_rows = deposit_attempted_rows or set()
    
    # Read the original CSV file
    original_df = pd.read_csv(bank_statement_path)
    
    # Calculate combined sets
    all_matched_rows = card_matched_rows | deposit_matched_rows
    all_attempted_rows = card_attempted_rows | deposit_attempted_rows
    overlap_matched = card_matched_rows & deposit_matched_rows
    failed_attempts = all_attempted_rows - all_matched_rows
    
    # Write to Excel with highlighting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        original_df.to_excel(writer, sheet_name='Bank Statement', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Bank Statement']
        
        # Define highlight colors
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')    # Card matches
        blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')     # Deposit matches
        purple_fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid')   # Both matched
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')     # Attempted but failed
        
        # Apply highlighting
        for row_idx in range(len(original_df)):
            excel_row = row_idx + 2  # +2 because Excel is 1-based and has header
            bank_row = excel_row  # bank_row is same as excel_row for highlighting
            
            # Determine highlight color based on priority: overlap > individual matches > failed attempts
            fill_color = None
            if bank_row in overlap_matched:
                # Purple for rows matched by both systems
                fill_color = purple_fill
            elif bank_row in card_matched_rows:
                # Green for card matches only
                fill_color = green_fill
            elif bank_row in deposit_matched_rows:
                # Blue for deposit matches only
                fill_color = blue_fill
            elif bank_row in failed_attempts:
                # Red for attempted but failed
                fill_color = red_fill
            
            # Apply the color if determined
            if fill_color:
                for col in range(1, len(original_df.columns) + 1):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.fill = fill_color
                    
                    # Add comment to the first cell of each matched row
                    if col == 1:
                        comment_text = None
                        
                        # Determine which type of match this is and get appropriate details
                        if bank_row in overlap_matched:
                            # Both card and deposit matched - combine information
                            card_details = card_transaction_details.get(bank_row) if card_transaction_details else None
                            deposit_details = deposit_transaction_details.get(bank_row) if deposit_transaction_details else None
                            comment_text = _create_combined_transaction_comment(
                                bank_row, card_details, deposit_details, 
                                card_match_type_info.get(bank_row) if card_match_type_info else None,
                                deposit_match_type_info.get(bank_row) if deposit_match_type_info else None
                            )
                        elif bank_row in card_matched_rows and card_transaction_details:
                            # Card match only
                            comment_text = _create_transaction_comment(
                                bank_row, card_transaction_details[bank_row],
                                None, card_match_type_info.get(bank_row) if card_match_type_info else None
                            )
                        elif bank_row in deposit_matched_rows and deposit_transaction_details:
                            # Deposit match only
                            comment_text = _create_transaction_comment(
                                bank_row, deposit_transaction_details[bank_row],
                                None, deposit_match_type_info.get(bank_row) if deposit_match_type_info else None
                            )
                        elif card_unmatched_transactions and bank_row in card_unmatched_transactions:
                            # Card unmatched transaction
                            comment_text = _create_unmatched_transaction_comment(
                                bank_row, card_unmatched_transactions[bank_row]
                            )
                        elif deposit_unmatched_transactions and bank_row in deposit_unmatched_transactions:
                            # Deposit unmatched transaction
                            comment_text = _create_unmatched_transaction_comment(
                                bank_row, deposit_unmatched_transactions[bank_row]
                            )
                        elif gc_transactions and bank_row in gc_transactions:
                            # GC transaction
                            comment_text = _create_gc_transaction_comment(
                                bank_row, gc_transactions[bank_row]
                            )
                        
                        if comment_text:
                            cell.comment = Comment(comment_text, "Bank Reconciliation System")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✓ Created combined highlighted bank statement: {output_path}")
    print("  Color legend:")
    print("    Green = Matched by credit card matching only")
    print("    Blue = Matched by deposit matching only") 
    print("    Purple = Matched by both credit card and deposit matching")
    print("    Red = Attempted but failed to match")
    print("    No color = Not processed for matching")
    print(f"  - Total rows: {len(original_df)}")
    print(f"  - Card matched: {len(card_matched_rows)}")
    print(f"  - Deposit matched: {len(deposit_matched_rows)}")
    print(f"  - Overlap (both): {len(overlap_matched)}")
    print(f"  - Total unique matched: {len(all_matched_rows)}")
    print(f"  - Failed match attempts: {len(failed_attempts)}")

def create_highlighted_card_summary(card_summary_path: str, matched_dates_and_types: dict,
                                  output_path: str = 'card_summary_highlighted.xlsx',
                                  skiprows=[0, 1, 3, 34], differences_info: dict = None,
                                  unmatched_info: dict = None, unmatched_bank_by_type: dict = None):
    """
    Create a highlighted copy of the card summary with matched cells highlighted and unmatched cells commented.
    
    Args:
        card_summary_path (str): Path to original card summary Excel
        matched_dates_and_types (dict): Dict with dates as keys and matched card types as values
        output_path (str): Output path for highlighted Excel file
        skiprows (list): Row indices that were skipped in original processing
        differences_info (dict): Optional dict with (date, card_type) tuples as keys and differences as values
        unmatched_info (dict): Dict with information about unmatched cells
        unmatched_bank_by_type (dict): Dict with unmatched bank transactions by card type, now includes date filter info
    """
    from dict import mapping
    
    # Copy the original Excel file
    shutil.copy2(card_summary_path, output_path)
    
    # Load the workbook
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    
    # Define highlight colors
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    # Read the data to understand the structure
    card_summary_df = pd.read_excel(card_summary_path, skiprows=skiprows)
    card_summary_df['Date'] = pd.to_datetime(card_summary_df['Date'])
    
    # Create a mapping of dates to their Excel row numbers
    excel_row_mapping = {}
    data_row_start = 5
    
    for idx, row in card_summary_df.iterrows():
        excel_row = data_row_start + idx
        if excel_row >= 34:
            excel_row += 1
        excel_row_mapping[row['Date']] = excel_row
    
    # Find column indices for card types
    header_row = 3
    column_mapping = {}
    
    for col_idx in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col_idx).value
        if cell_value:
            column_mapping[str(cell_value).strip()] = col_idx
    
    # Track statistics
    matched_cells_count = 0
    unmatched_cells_count = 0
    
    # Find the total row first
    total_row = None
    for row in range(worksheet.max_row, 0, -1):
        cell_value = worksheet.cell(row=row, column=1).value
        if cell_value and 'Total' in str(cell_value):
            total_row = row
            break

    # Process all data rows
    for _, row in card_summary_df.iterrows():
        date = row['Date']
        excel_row = excel_row_mapping[date]
        
        # Skip highlighting if this is the total row
        if total_row and excel_row == total_row:
            continue
        
        # Check each card type column
        for card_type in column_mapping.keys():
            if card_type in ['Date', 'Total', 'Visa & MC'] or card_type.startswith('Unnamed'):
                continue
                
            if card_type in column_mapping and card_type in row.index:
                expected_amount = row[card_type]
                if pd.notna(expected_amount) and expected_amount != 0:
                    col_idx = column_mapping[card_type]
                    cell = worksheet.cell(row=excel_row, column=col_idx)
                    
                    # Check if this cell was matched
                    if matched_dates_and_types and date in matched_dates_and_types and card_type in matched_dates_and_types[date]:
                        # Matched cell - color green
                        cell.fill = green_fill
                        matched_cells_count += 1
                    elif unmatched_info and (date, card_type) in unmatched_info:
                        # Unmatched cell - color red and add comment
                        cell.fill = red_fill
                        unmatched_cells_count += 1
                        
                        # Add comment with what was actually found
                        info = unmatched_info[(date, card_type)]
                        found_count = info.get('found_transactions', 0)
                        total_found = info.get('total_found', 0)
                        expected = info.get('expected', expected_amount)
                        difference = total_found - expected
                        
                        comment_lines = []
                        comment_lines.append(f"Expected: ${expected:,.2f}")
                        comment_lines.append(f"Found transactions: ${total_found:,.2f}")
                        comment_lines.append(f"Difference: ${difference:,.2f}")
                        comment_lines.append(f"Unmatched transactions available: {found_count}")
                        if info.get('reason'):
                            comment_lines.append(f"Reason: {info['reason']}")
                        comment_lines.append("\nNote: 'Found' amount excludes transactions already matched elsewhere")
                        
                        comment_text = "\n".join(comment_lines)
                        cell.comment = Comment(comment_text, "Matching System")
    
    # Add unmatched BANK STATEMENT totals to the Total row with date filtering info
    if total_row and unmatched_bank_by_type:
        # Add comments to total row cells
        for card_type, col_idx in column_mapping.items():
            if card_type in unmatched_bank_by_type and card_type not in ['Date', 'Total', 'Visa & MC']:
                cell = worksheet.cell(row=total_row, column=col_idx)
                totals = unmatched_bank_by_type[card_type]
                
                comment_lines = []
                comment_lines.append(f"Unmatched bank transactions:")
                
                # Add date filter information if applicable
                if totals.get('date_filter_applied', False) and totals.get('first_matched_date'):
                    first_date = totals['first_matched_date']
                    comment_lines.append(f"Counting from: {first_date.strftime('%Y-%m-%d')} (first match)")
                elif not totals.get('date_filter_applied', False):
                    comment_lines.append(f"All transactions (no matches found)")
                
                comment_lines.append(f"Total: ${totals['total']:,.2f}")
                comment_lines.append(f"Count: {totals['count']} transactions")
                comment_lines.append(f"\nBank rows: {', '.join(map(str, totals['rows'][:10]))}")
                if totals['count'] > 10:
                    comment_lines.append(f"... and {totals['count'] - 10} more")
                
                comment_text = "\n".join(comment_lines)
                cell.comment = Comment(comment_text, "Bank Statement Summary")
    
    # Save the workbook
    workbook.save(output_path)
    workbook.close()
    
    print(f"✓ Created highlighted card summary: {output_path}")
    print(f"  - Total dates: {len(card_summary_df)}")
    print(f"  - Matched cells (green): {matched_cells_count}")
    print(f"  - Unmatched cells with comments (red): {unmatched_cells_count}")
    if total_row and unmatched_bank_by_type:
        print(f"  - Total row comments added for unmatched bank transactions: {len([k for k in unmatched_bank_by_type.keys() if k in column_mapping])}")

def extract_matched_info_from_results(results: dict) -> tuple:
    """
    Extract matched bank rows and matched dates/card types from results.
    Also extracts unmatched information for highlighting.
    
    Returns:
        tuple: (set of matched bank rows, dict of matched dates and card types, 
                dict of differences by bank row, dict of differences by date/card_type,
                dict of unmatched info)
    """
    matched_bank_rows = set()
    matched_dates_and_types = {}
    differences_by_row = {}
    differences_by_date_type = {}
    unmatched_info = {}
    
    for date, date_results in results.items():
        # Skip metadata keys that start with underscore
        if isinstance(date, str) and date.startswith('_'):
            continue
        
        # Also skip if date_results is not a dictionary with expected structure
        if not isinstance(date_results, dict) or 'matches_by_card_type' not in date_results:
            continue
            
        matched_types = []
        
        # Process matched transactions
        for card_type, match_info in date_results['matches_by_card_type'].items():
            matched_types.append(card_type)
            matched_bank_rows.update(match_info['bank_rows'])
            
            diff = match_info.get('difference', 0)
            if abs(diff) > 0.01:
                for bank_row in match_info['bank_rows']:
                    differences_by_row[bank_row] = diff
                differences_by_date_type[(date, card_type)] = diff
        
        if matched_types:
            matched_dates_and_types[date] = matched_types
        
        # Process unmatched transactions
        for card_type, unmatch_info in date_results['unmatched_by_card_type'].items():
            unmatched_info[(date, card_type)] = unmatch_info
    
    print(f"\n=== Match Summary ===")
    print(f"Matched cells: {sum(len(types) for types in matched_dates_and_types.values())}")
    print(f"Unmatched cells: {len(unmatched_info)}")
    
    return matched_bank_rows, matched_dates_and_types, differences_by_row, differences_by_date_type, unmatched_info

def extract_transaction_details_for_comments(results: dict, bank_statement: pd.DataFrame) -> tuple:
    """
    Extract transaction details and match type information for adding comments to bank statement.
    
    Args:
        results (dict): Matching results from the transaction matcher
        bank_statement (pd.DataFrame): Bank statement dataframe
        
    Returns:
        tuple: (dict of transaction details by bank row, dict of match types by bank row)
    """
    transaction_details = {}
    match_type_info = {}
    
    for date, date_results in results.items():
        # Skip metadata keys that start with underscore
        if isinstance(date, str) and date.startswith('_'):
            continue
        
        # Also skip if date_results is not a dictionary with expected structure
        if not isinstance(date_results, dict) or 'matches_by_card_type' not in date_results:
            continue
        
        # Process matched transactions
        for card_type, match_info in date_results['matches_by_card_type'].items():
            match_type = match_info.get('match_type', 'unknown')
            
            # Get transaction details for each matched bank row
            for bank_row in match_info['bank_rows']:
                # Find the transaction in the bank statement
                if bank_row <= len(bank_statement) + 1:  # +1 for header row
                    bank_idx = bank_row - 2  # Convert to 0-based index
                    if 0 <= bank_idx < len(bank_statement):
                        transaction = bank_statement.iloc[bank_idx]
                        
                        # Create transaction details for comment
                        transaction_details[bank_row] = {
                            'Date': transaction.get('Date', ''),
                            'Description': transaction.get('Description', ''),
                            'Amount': transaction.get('Amount', 0),
                            'Card_Type': transaction.get('Card_Type', ''),
                            'Transaction_Type': transaction.get('Transaction_Type', ''),
                            'Expected_Amount': match_info.get('expected', 0),
                            'Actual_Total': match_info.get('actual_total', 0)
                        }
                        
                        # Store match type information
                        match_type_info[bank_row] = match_type
    
    return transaction_details, match_type_info

def extract_unmatched_transactions_for_comments(results: dict, bank_statement: pd.DataFrame) -> dict:
    """
    Extract unmatched transaction information for adding comments to bank statement.
    
    Args:
        results (dict): Matching results from the transaction matcher
        bank_statement (pd.DataFrame): Bank statement dataframe
        
    Returns:
        dict: Dictionary mapping bank rows to unmatched transaction details
    """
    unmatched_transactions = {}
    
    for date, date_results in results.items():
        # Skip metadata keys that start with underscore
        if isinstance(date, str) and date.startswith('_'):
            continue
        
        # Also skip if date_results is not a dictionary with expected structure
        if not isinstance(date_results, dict) or 'unmatched_by_card_type' not in date_results:
            continue
        
        # Process unmatched transactions
        for card_type, unmatch_info in date_results['unmatched_by_card_type'].items():
            # Get bank rows that were attempted but not matched
            if 'bank_rows' in unmatch_info and unmatch_info['bank_rows']:
                for bank_row in unmatch_info['bank_rows']:
                    # Find the transaction in the bank statement
                    if bank_row <= len(bank_statement) + 1:  # +1 for header row
                        bank_idx = bank_row - 2  # Convert to 0-based index
                        if 0 <= bank_idx < len(bank_statement):
                            transaction = bank_statement.iloc[bank_idx]
                            
                            # Create unmatched transaction details for comment
                            unmatched_transactions[bank_row] = {
                                'Date': transaction.get('Date', ''),
                                'Description': transaction.get('Description', ''),
                                'Amount': transaction.get('Amount', 0),
                                'Card_Type': transaction.get('Card_Type', ''),
                                'Transaction_Type': transaction.get('Transaction_Type', ''),
                                'reason': unmatch_info.get('reason', 'No match found'),
                                'filters_tried': unmatch_info.get('filters_tried', []),
                                'expected': unmatch_info.get('expected', 0),
                                'total_found': unmatch_info.get('total_found', 0),
                                'found_transactions': unmatch_info.get('found_transactions', 0)
                            }
    
    return unmatched_transactions

def extract_gc_transactions_for_comments(results: dict, bank_statement: pd.DataFrame) -> dict:
    """
    Extract GC (Cash & Check) transaction information for adding comments to bank statement.
    
    Args:
        results (dict): Matching results from the transaction matcher
        bank_statement (pd.DataFrame): Bank statement dataframe
        
    Returns:
        dict: Dictionary mapping bank rows to GC transaction details
    """
    gc_transactions = {}
    
    # Look for GC 1416 transactions in the bank statement
    gc_transactions_df = bank_statement[
        (bank_statement['Description'].str.contains('GC 1416', case=False, na=False)) |
        (bank_statement['Description'].str.contains('CASH/CHECK', case=False, na=False)) |
        (bank_statement['Description'].str.contains('Cash/Check', case=False, na=False))
    ]
    
    for _, transaction in gc_transactions_df.iterrows():
        bank_row = transaction.get('Bank_Row_Number', 0)
        if bank_row > 0:
            # Create GC transaction details for comment
            gc_transactions[bank_row] = {
                'Date': transaction.get('Date', ''),
                'Description': transaction.get('Description', ''),
                'Amount': transaction.get('Amount', 0),
                'Transaction_Type': transaction.get('Transaction_Type', ''),
                'match_type': 'gc_1416_deposit',
                'allocation': {
                    'cash_amount': 0,  # Will be filled by deposit matching results
                    'check_amount': 0  # Will be filled by deposit matching results
                }
            }
    
    # Try to get allocation information from deposit matching results
    for date, date_results in results.items():
        if isinstance(date, str) and date.startswith('_'):
            continue
        
        if isinstance(date_results, dict) and 'matches_by_type' in date_results:
            # This is deposit matching results
            for deposit_type, match_info in date_results['matches_by_type'].items():
                if 'bank_rows' in match_info:
                    for bank_row in match_info['bank_rows']:
                        if bank_row in gc_transactions:
                            # Update allocation information
                            if deposit_type == 'Cash':
                                gc_transactions[bank_row]['allocation']['cash_amount'] = match_info.get('actual_total', 0)
                            elif deposit_type == 'Check':
                                gc_transactions[bank_row]['allocation']['check_amount'] = match_info.get('actual_total', 0)
                            
                            # Add deposit date
                            gc_transactions[bank_row]['deposit_date'] = date
    
    return gc_transactions