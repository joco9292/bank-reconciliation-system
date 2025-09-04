import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill

# Import preprocessor functions
from preprocess_bank_statement import preprocess_bank_statement
from preprocess_card_summary import preprocess_card_summary_dynamic, create_highlighted_card_summary_dynamic

# Import helper functions and classes
from matching_helpers import (
    identify_card_type, 
    TransactionMatcher,
    filter_by_amount_range,
    filter_split_transactions
)

# Import the new highlighting functions
from highlighting_functions import (
    create_highlighted_bank_statement,
    extract_matched_info_from_results
)

def prepare_data_for_matching(card_summary: pd.DataFrame, bank_statement: pd.DataFrame) -> tuple:
    """
    Add additional columns needed for matching process.
    """
    # Add row numbers to bank statement (1-based to match Excel)
    bank_statement['Bank_Row_Number'] = range(2, len(bank_statement) + 2)
    
    # Identify card type for each bank transaction
    bank_statement['Card_Type'] = bank_statement['Description'].apply(identify_card_type)
    
    return card_summary, bank_statement

def find_first_matched_date(matched_bank_rows: set, bank_statement: pd.DataFrame) -> pd.Timestamp:
    """
    Find the date of the earliest matched transaction.
    """
    if not matched_bank_rows:
        return None
    
    matched_transactions = bank_statement[bank_statement['Bank_Row_Number'].isin(matched_bank_rows)]
    if matched_transactions.empty:
        return None
    
    return matched_transactions['Date'].min()

def calculate_total_discrepancies_by_card_type(results: dict, bank_statement: pd.DataFrame, 
                                              matched_bank_rows: set) -> dict:
    """
    Calculate total discrepancies by comparing:
    - Unmatched bank transactions (after first match) 
    - Unmatched card summary expectations
    
    Returns:
        dict: Card type -> net discrepancy (positive = bank has more, negative = card summary expects more)
    """
    discrepancies_by_type = {}
    
    # Find the first matched transaction date
    first_matched_date = find_first_matched_date(matched_bank_rows, bank_statement)
    
    # Step 1: Add up all unmatched BANK transactions (after first match)
    if first_matched_date:
        unmatched_bank_df = bank_statement[
            (~bank_statement['Bank_Row_Number'].isin(matched_bank_rows)) &
            (bank_statement['Date'] > first_matched_date)
        ]
    else:
        # If no matches at all, count all bank transactions
        unmatched_bank_df = bank_statement[~bank_statement['Bank_Row_Number'].isin(matched_bank_rows)]
    
    # Sum unmatched bank amounts by card type (these are EXTRA transactions we found)
    for card_type in unmatched_bank_df['Card_Type'].unique():
        if card_type != 'Unknown':
            type_df = unmatched_bank_df[unmatched_bank_df['Card_Type'] == card_type]
            if card_type not in discrepancies_by_type:
                discrepancies_by_type[card_type] = 0
            # Add what we found but didn't expect
            discrepancies_by_type[card_type] += type_df['Amount'].sum()
    
    # Step 2: Subtract all unmatched CARD SUMMARY expectations
    for date, date_results in results.items():
        for card_type, unmatch_info in date_results['unmatched_by_card_type'].items():
            if card_type not in discrepancies_by_type:
                discrepancies_by_type[card_type] = 0
            # Subtract what we expected but didn't find
            discrepancies_by_type[card_type] -= unmatch_info.get('expected', 0)
    
    return discrepancies_by_type, first_matched_date

def generate_enhanced_report(results: dict, bank_statement: pd.DataFrame, 
                           output_path: str = 'matching_report_enhanced.xlsx'):
    """
    Generate report with bank row references and formatting.
    """
    summary_data = []
    detailed_matches = []
    unmatched_data = []
    
    for date, date_results in results.items():
        # Matched transactions
        for card_type, match_info in date_results['matches_by_card_type'].items():
            summary_data.append({
                'Date': date,
                'Card_Type': card_type,
                'Expected_Amount': match_info['expected'],
                'Match_Type': match_info['match_type'],
                'Status': 'Matched',
                'Bank_Rows': ', '.join(map(str, match_info['bank_rows'])),
                'Transaction_Count': len(match_info['transactions'])
            })
            
            # Detailed match information
            for trans in match_info['transactions']:
                detailed_matches.append({
                    'Match_Date': date,
                    'Card_Type': card_type,
                    'Match_Type': match_info['match_type'],
                    'Bank_Row': trans['Bank_Row_Number'],
                    'Transaction_Date': trans['Date'],
                    'Description': trans['Description'],
                    'Amount': trans['Amount']
                })
        
        # Unmatched transactions
        for card_type, unmatch_info in date_results['unmatched_by_card_type'].items():
            summary_data.append({
                'Date': date,
                'Card_Type': card_type,
                'Expected_Amount': unmatch_info['expected'],
                'Match_Type': 'Unmatched',
                'Status': 'Unmatched',
                'Bank_Rows': ', '.join(map(str, unmatch_info.get('bank_rows', []))),
                'Reason': unmatch_info['reason']
            })
            
            unmatched_data.append({
                'Date': date,
                'Card_Type': card_type,
                'Expected_Amount': unmatch_info['expected'],
                'Found_Transactions': unmatch_info.get('found_transactions', 0),
                'Total_Found': unmatch_info.get('total_found', 0),
                'Reason': unmatch_info['reason']
            })
    
    # Create Excel report with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Add formatting to highlight matched/unmatched
            worksheet = writer.sheets['Summary']
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            for idx, row in enumerate(summary_df.itertuples(), start=2):
                if row.Status == 'Matched':
                    worksheet[f'E{idx}'].fill = green_fill
                else:
                    worksheet[f'E{idx}'].fill = red_fill
        
        if detailed_matches:
            pd.DataFrame(detailed_matches).to_excel(writer, sheet_name='Matched_Details', index=False)
        
        if unmatched_data:
            pd.DataFrame(unmatched_data).to_excel(writer, sheet_name='Unmatched', index=False)
        
        # Add bank statement reference sheet
        bank_ref = bank_statement[['Bank_Row_Number', 'Date', 'Description', 'Amount', 'Card_Type']]
        bank_ref.to_excel(writer, sheet_name='Bank_Statement_Reference', index=False)

def print_matching_summary(results: dict, discrepancies_by_type: dict, first_matched_date):
    """
    Print a summary of matching results.
    """
    total_matches = 0
    total_unmatched = 0
    matches_by_type = {}
    
    for date, date_results in results.items():
        total_matches += len(date_results['matches_by_card_type'])
        total_unmatched += len(date_results['unmatched_by_card_type'])
        
        for match_info in date_results['matches_by_card_type'].values():
            match_type = match_info['match_type']
            matches_by_type[match_type] = matches_by_type.get(match_type, 0) + 1
    
    print("\n=== MATCHING SUMMARY ===")
    print(f"Total matched: {total_matches}")
    print(f"Total unmatched: {total_unmatched}")
    print(f"\nMatches by type:")
    for match_type, count in matches_by_type.items():
        print(f"  - {match_type}: {count}")
    
    # Print discrepancies by card type
    print("\n=== NET DISCREPANCIES BY CARD TYPE ===")
    if first_matched_date:
        print(f"(Calculated from {first_matched_date.strftime('%Y-%m-%d')} onwards)\n")
    
    total_discrepancy = 0
    for card_type, disc in sorted(discrepancies_by_type.items()):
        if abs(disc) > 0.01:  # Only show non-zero discrepancies
            if disc > 0:
                print(f"{card_type}: +${disc:,.2f} (bank has more than expected)")
            else:
                print(f"{card_type}: -${abs(disc):,.2f} (bank has less than expected)")
            total_discrepancy += disc
    
    print(f"\nTotal net discrepancy: ${total_discrepancy:,.2f}")
    if total_discrepancy > 0:
        print("(Positive = bank statement total exceeds card summary expectations)")
    elif total_discrepancy < 0:
        print("(Negative = card summary expects more than bank statement shows)")

if __name__ == "__main__":
    print("=== Transaction Matching System ===\n")
    
    # Store file paths for later use
    BANK_STATEMENT_PATH = 'june 2025 bank statement.CSV'
    CARD_SUMMARY_PATH = 'XYZ Storage Laird - CreditCardSummary - 07-01-2025 - 07-31-2025 (3).xlsx'
    
    # Step 1: Load and preprocess data
    print("Step 1: Loading and preprocessing data...")
    try:
        bank_statement = preprocess_bank_statement(BANK_STATEMENT_PATH)
        print(f"✓ Loaded bank statement: {len(bank_statement)} transactions")
    except Exception as e:
        print(f"✗ Error loading bank statement: {e}")
        exit(1)
    
    try:
        card_summary, structure_info = preprocess_card_summary_dynamic(CARD_SUMMARY_PATH)
        print(f"✓ Loaded card summary: {len(card_summary)} days")
    except Exception as e:
        print(f"✗ Error loading card summary: {e}")
        exit(1)
    
    # Step 2: Prepare data for matching
    print("\nStep 2: Preparing data for matching...")
    card_summary, bank_statement = prepare_data_for_matching(card_summary, bank_statement)
    
    # Show card type distribution
    print("\nCard types identified in bank statement:")
    print(bank_statement['Card_Type'].value_counts())
    
    # Step 3: Create matcher with card-type-specific filters
    print("\nStep 3: Setting up transaction matcher with card-type-specific filters...")
    matcher = TransactionMatcher()
    
    # OPTION 1: Add filters to specific card types only
    # This adds amount_range ONLY for Amex transactions
    matcher.add_filter('amount_range', filter_by_amount_range, card_types=['Amex'])
    
    # OPTION 2: You can also add filters to multiple specific card types
    # matcher.add_filter('split_transactions', filter_split_transactions, 
    #                   card_types=['Discover', 'Other Cards'])
    
    # OPTION 3: Add a filter to ALL card types (default behavior)
    # matcher.add_filter('new_filter', new_filter_function)  # No card_types param = applies to all
    
    # OPTION 4: You can also modify the class initialization directly to set up
    # card-specific filters (see the TransactionMatcher.__init__ method)
    
    # Alternative approach: Create a custom matcher with specific configuration
    # custom_matcher = create_custom_matcher()
    
    # Step 4: Run matching algorithm with verbose output to see which filters are used
    print("\nStep 4: Running matching algorithm...")
    print("Note: amount_range filter will ONLY be applied to Amex transactions\n")
    
    # Set verbose=True to see which filters are applied to each card type
    results = matcher.match_transactions(card_summary, bank_statement, 
                                        forward_days=3, verbose=False)
    
    # Rest of your main.py code remains the same...
    # Step 5: Extract matched information
    matched_bank_rows, matched_dates_and_types, differences_by_row, differences_by_date_type, unmatched_info = extract_matched_info_from_results(results)
    
    # Step 6: Calculate comprehensive discrepancies from bank statement perspective
    print("\nStep 5: Calculating net discrepancies...")
    discrepancies_by_type, first_matched_date = calculate_total_discrepancies_by_card_type(
        results, bank_statement, matched_bank_rows
    )
    
    # Step 7: Generate summary and report
    print("\nStep 6: Generating results...")
    print_matching_summary(results, discrepancies_by_type, first_matched_date)
    
    # Generate Excel report
    output_filename = 'matching_report_enhanced.xlsx'
    generate_enhanced_report(results, bank_statement, output_filename)
    print(f"\n✓ Detailed report saved to: {output_filename}")
    
    # Step 8: Create highlighted copies of original files
    print("\nStep 7: Creating highlighted copies of original files...")
    
    # Create highlighted bank statement
    create_highlighted_bank_statement(
        bank_statement_path=BANK_STATEMENT_PATH,
        matched_bank_rows=matched_bank_rows,
        output_path='bank_statement_highlighted.xlsx',
        differences_by_row=differences_by_row
    )
    
    # Create highlighted card summary with comprehensive discrepancies
    create_highlighted_card_summary_dynamic(
        card_summary_path=CARD_SUMMARY_PATH,
        matched_dates_and_types=matched_dates_and_types,
        output_path='card_summary_highlighted.xlsx',
        differences_info=differences_by_date_type,
        unmatched_info=unmatched_info,
        differences_by_card_type=discrepancies_by_type
    )
    
    # Final summary
    print("\n=== PROCESS COMPLETE ===")
    print("\nGenerated files:")
    print("  1. matching_report_enhanced.xlsx - Detailed matching report")
    print("  2. bank_statement_highlighted.xlsx - Bank statement with matched rows highlighted")
    print("  3. card_summary_highlighted.xlsx - Card summary with matched cells highlighted")
    
    if first_matched_date:
        print(f"\nNote: Discrepancies calculated from {first_matched_date.strftime('%Y-%m-%d')} onwards")
        print("(Earlier bank transactions excluded as they belong to previous month)")


# Optional: Create a function to build a custom matcher with complex configurations
def create_custom_matcher():
    """
    Create a TransactionMatcher with a complex custom configuration.
    """
    matcher = TransactionMatcher()
    
    # Configure different filters for different card types
    
    # Amex: Use looser matching including amount range
    matcher.add_filter('amount_range', 
                      lambda trans, amt: filter_by_amount_range(trans, amt, percentage_tolerance=0.10),
                      card_types=['Amex'])
    
    # Discover: Try split transactions
    matcher.add_filter('split_transactions', filter_split_transactions, 
                      card_types=['Discover'])
    
    # Debit cards: Use stricter matching (default filters only)
    # No additional filters needed
    
    # Master Card: Use a tighter amount range
    matcher.add_filter('amount_range_tight',
                      lambda trans, amt: filter_by_amount_range(trans, amt, percentage_tolerance=0.02),
                      card_types=['Master Card'])
    
    return matcher